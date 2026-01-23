#!/usr/bin/env python3
"""
Import inventory from the cleaned XLSX file into the database.

This script:
1. Reads the 'All Inventory' sheet from the XLSX
2. Creates/finds Brand, ProductLine, Checklist records
3. Creates Inventory records with full cost breakdown

Usage:
    python import_inventory_xlsx.py <xlsx_file_path> [--execute]

    Without --execute, runs in preview mode (no DB changes).
"""

import asyncio
import sys
import os
from decimal import Decimal
from collections import defaultdict

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app.database import AsyncSessionLocal
from app.models import Brand, ProductLine, Checklist, Inventory
from app.models.card_types import CardBaseType, Parallel, ParallelCategory


async def get_or_create_brand(db: AsyncSession, name: str, cache: dict) -> "Brand":
    """Get or create a Brand record."""
    name_clean = name.strip()
    if name_clean in cache:
        return cache[name_clean]

    result = await db.execute(select(Brand).where(Brand.name == name_clean))
    brand = result.scalar_one_or_none()

    if not brand:
        slug = name_clean.lower().replace(' ', '-')
        brand = Brand(name=name_clean, slug=slug)
        db.add(brand)
        await db.flush()

    cache[name_clean] = brand
    return brand


async def get_or_create_product_line(
    db: AsyncSession, brand: "Brand", year: int, cache: dict
) -> "ProductLine":
    """Get or create a ProductLine record."""
    key = (str(brand.id), year)
    if key in cache:
        return cache[key]

    result = await db.execute(
        select(ProductLine).where(
            ProductLine.brand_id == brand.id,
            ProductLine.year == year,
            ProductLine.name == brand.name,
        )
    )
    pl = result.scalar_one_or_none()

    if not pl:
        pl = ProductLine(
            brand_id=brand.id,
            name=brand.name,
            year=year,
            sport="Baseball",
        )
        db.add(pl)
        await db.flush()

    cache[key] = pl
    return pl


async def get_or_create_base_type(db: AsyncSession, name: str, cache: dict) -> "CardBaseType":
    """Get or create a CardBaseType record."""
    if name in cache:
        return cache[name]

    result = await db.execute(select(CardBaseType).where(CardBaseType.name == name))
    bt = result.scalar_one_or_none()

    if not bt:
        bt = CardBaseType(name=name, sort_order=len(cache))
        db.add(bt)
        await db.flush()

    cache[name] = bt
    return bt


async def get_or_create_parallel(db: AsyncSession, name: str, category_id, cache: dict) -> "Parallel":
    """Get or create a Parallel record."""
    if name in cache:
        return cache[name]

    result = await db.execute(select(Parallel).where(Parallel.name == name))
    p = result.scalar_one_or_none()

    if not p:
        p = Parallel(
            name=name,
            short_name=name[:20],
            category_id=category_id,
            sort_order=len(cache),
        )
        db.add(p)
        await db.flush()

    cache[name] = p
    return p


async def get_or_create_parallel_category(db: AsyncSession, cache: dict) -> "ParallelCategory":
    """Get or create a default ParallelCategory."""
    if "__default__" in cache:
        return cache["__default__"]

    result = await db.execute(select(ParallelCategory).where(ParallelCategory.name == "Imported"))
    cat = result.scalar_one_or_none()

    if not cat:
        cat = ParallelCategory(name="Imported", sort_order=99)
        db.add(cat)
        await db.flush()

    cache["__default__"] = cat
    return cat


async def get_or_create_checklist(
    db: AsyncSession,
    player_name: str,
    product_line: "ProductLine",
    base_type: "CardBaseType",
    cache: dict,
) -> "Checklist":
    """Get or create a Checklist entry for this player/product combo."""
    key = (player_name, str(product_line.id), str(base_type.id))
    if key in cache:
        return cache[key]

    # Use a synthetic card number based on player name
    card_number = f"IMP-{player_name[:3].upper()}"

    result = await db.execute(
        select(Checklist).where(
            Checklist.product_line_id == product_line.id,
            Checklist.player_name_raw == player_name,
            Checklist.set_name == f"Imported-{base_type.name}",
        )
    )
    cl = result.scalar_one_or_none()

    if not cl:
        cl = Checklist(
            product_line_id=product_line.id,
            card_number=card_number,
            player_name_raw=player_name,
            base_type_id=base_type.id,
            set_name=f"Imported-{base_type.name}",
        )
        db.add(cl)
        await db.flush()

    cache[key] = cl
    return cl


def read_xlsx(file_path: str) -> list[dict]:
    """Read the All Inventory sheet and return list of row dicts."""
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb['All Inventory']

    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        headers.append(cell.value)

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = val
        rows.append(row_dict)

    wb.close()
    return rows


def preview(rows: list[dict]):
    """Print a preview of what will be imported."""
    total_qty = sum(r.get('quantity', 0) or 0 for r in rows)
    total_cost = sum(Decimal(str(r.get('total_cost', 0) or 0)) for r in rows)

    # Count by status
    unsigned_raw = [r for r in rows if not r.get('is_signed') and not r.get('is_slabbed')]
    signed_raw = [r for r in rows if r.get('is_signed') and not r.get('is_slabbed')]
    slabbed = [r for r in rows if r.get('is_slabbed')]

    unsigned_qty = sum(r.get('quantity', 0) or 0 for r in unsigned_raw)
    signed_qty = sum(r.get('quantity', 0) or 0 for r in signed_raw)
    slabbed_qty = sum(r.get('quantity', 0) or 0 for r in slabbed)

    # Unique values
    brands = set(r.get('brand', '').strip() for r in rows if r.get('brand'))
    players = set(r.get('player_name', '') for r in rows if r.get('player_name'))

    print(f"\n{'='*60}")
    print("IMPORT PREVIEW")
    print(f"{'='*60}")
    print(f"  Total rows:          {len(rows)}")
    print(f"  Total cards:         {total_qty:,}")
    print(f"  Total investment:    ${total_cost:,.2f}")
    print(f"  Unique players:      {len(players)}")
    print(f"  Unique brands:       {brands}")
    print()
    print("  BY STATUS:")
    print(f"    Raw Unsigned:      {unsigned_qty:,} cards ({len(unsigned_raw)} rows)")
    print(f"    Raw Signed:        {signed_qty:,} cards ({len(signed_raw)} rows)")
    print(f"    Slabbed:           {slabbed_qty:,} cards ({len(slabbed)} rows)")
    print()
    print("  COST BREAKDOWN:")
    card_cost = sum(Decimal(str(r.get('card_cost', 0) or 0)) for r in rows)
    sign_cost = sum(Decimal(str(r.get('signing_cost', 0) or 0)) for r in rows)
    grade_cost = sum(Decimal(str(r.get('grading_cost', 0) or 0)) for r in rows)
    print(f"    Card purchases:    ${card_cost:,.2f}")
    print(f"    Signing fees:      ${sign_cost:,.2f}")
    print(f"    Grading fees:      ${grade_cost:,.2f}")
    print()

    # Consigner breakdown
    consigners = defaultdict(int)
    for r in rows:
        if r.get('consigner'):
            consigners[r['consigner']] += r.get('quantity', 0) or 0
    if consigners:
        print("  CONSIGNERS:")
        for c, qty in sorted(consigners.items(), key=lambda x: -x[1]):
            print(f"    {c}: {qty} cards")

    print(f"\n{'='*60}\n")


async def execute_import(rows: list[dict]):
    """Execute the import into the database."""
    # Caches for get_or_create
    brand_cache = {}
    pl_cache = {}
    bt_cache = {}
    parallel_cache = {}
    cat_cache = {}
    checklist_cache = {}

    success_count = 0
    error_count = 0
    errors = []

    async with AsyncSessionLocal() as db:
        try:
            # Get/create default parallel category
            default_category = await get_or_create_parallel_category(db, cat_cache)

            for i, row in enumerate(rows):
                try:
                    player_name = row.get('player_name', '').strip()
                    brand_name = (row.get('brand', '') or '').strip()
                    year = int(row.get('year', 0) or 0)
                    base_type_name = (row.get('base_type', '') or 'Chrome').strip()
                    parallel_name = (row.get('parallel', '') or '').strip()
                    quantity = int(row.get('quantity', 0) or 0)
                    is_signed = bool(row.get('is_signed'))
                    is_slabbed = bool(row.get('is_slabbed'))
                    grade_company = (row.get('grade_company', '') or '').strip() or None
                    consigner = (row.get('consigner', '') or '').strip() or None
                    how_obtained = (row.get('how_obtained', '') or '').strip() or None
                    card_cost = Decimal(str(row.get('card_cost', 0) or 0))
                    signing_cost = Decimal(str(row.get('signing_cost', 0) or 0))
                    grading_cost = Decimal(str(row.get('grading_cost', 0) or 0))
                    total_cost = Decimal(str(row.get('total_cost', 0) or 0))

                    if not player_name or not brand_name or not year:
                        errors.append(f"Row {i+2}: Missing player_name, brand, or year")
                        error_count += 1
                        continue

                    # Get/create hierarchy
                    brand = await get_or_create_brand(db, brand_name, brand_cache)
                    product_line = await get_or_create_product_line(db, brand, year, pl_cache)
                    base_type = await get_or_create_base_type(db, base_type_name, bt_cache)

                    parallel = None
                    if parallel_name:
                        parallel = await get_or_create_parallel(
                            db, parallel_name, default_category.id, parallel_cache
                        )

                    checklist = await get_or_create_checklist(
                        db, player_name, product_line, base_type, checklist_cache
                    )

                    # Create inventory record
                    inv = Inventory(
                        item_type='card',
                        checklist_id=checklist.id,
                        base_type_id=base_type.id,
                        parallel_id=parallel.id if parallel else None,
                        quantity=quantity,
                        is_signed=is_signed,
                        is_slabbed=is_slabbed,
                        grade_company=grade_company,
                        raw_condition="NM",
                        card_cost=card_cost,
                        signing_cost=signing_cost,
                        grading_cost=grading_cost,
                        total_cost=total_cost,
                        consigner=consigner,
                        how_obtained=how_obtained,
                    )
                    db.add(inv)
                    success_count += 1

                    # Flush every 100 rows
                    if (i + 1) % 100 == 0:
                        await db.flush()
                        print(f"  Processed {i+1}/{len(rows)} rows...")

                except Exception as e:
                    error_count += 1
                    errors.append(f"Row {i+2}: {str(e)}")

            await db.commit()

        except Exception as e:
            await db.rollback()
            print(f"\nFATAL ERROR: {e}")
            raise

    print(f"\n{'='*60}")
    print("IMPORT COMPLETE")
    print(f"{'='*60}")
    print(f"  Success: {success_count}")
    print(f"  Errors:  {error_count}")
    if errors:
        print(f"\n  First 20 errors:")
        for err in errors[:20]:
            print(f"    {err}")
    print()


async def main():
    if len(sys.argv) < 2:
        print("Usage: python import_inventory_xlsx.py <xlsx_file_path> [--execute]")
        print("\nOptions:")
        print("  --execute    Actually run the import (default is preview only)")
        sys.exit(1)

    file_path = sys.argv[1]
    execute = '--execute' in sys.argv

    if not os.path.exists(file_path):
        print(f"Error: File not found: {file_path}")
        sys.exit(1)

    print(f"Reading: {file_path}")
    rows = read_xlsx(file_path)
    print(f"Found {len(rows)} rows")

    preview(rows)

    if execute:
        print("EXECUTING IMPORT...")
        await execute_import(rows)
    else:
        print("Preview only. Use --execute to run the import.")


if __name__ == "__main__":
    asyncio.run(main())
